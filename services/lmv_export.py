"""
Export Live Master View data to an .xlsx file.

Column headers are overridden by the 'Main Column Name' rename map from
config_store: any displayed header that has a renamed entry is written using
the renamed name. Headers without an override keep their actual name.
"""

import io
import os


def apply_rename(headers: list, rename_map: dict) -> list:
    """Return headers with any actual→renamed overrides applied."""
    return [rename_map.get(h, h) for h in headers]


def export_xlsx(output_path: str, headers: list, rows: list,
                rename_map: dict = None) -> None:
    """Write *headers*/*rows* to *output_path* as a styled .xlsx workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if rename_map is None:
        from services import config_store
        rename_map = config_store.get_rename_map()

    out_headers = apply_rename(headers, rename_map)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LMV Export"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")

    for col_idx, header in enumerate(out_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    parent = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(parent, exist_ok=True)

    buf = io.BytesIO()
    wb.save(buf)
    with open(output_path, "wb") as f:
        f.write(buf.getvalue())
