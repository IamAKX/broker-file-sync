"""
Read broker export files and extract the configured columns.

Column specs (0-based indices):
  Sharekhan       (.xlsx/.xls): header at row 8 (idx 7), data from row 9
                  cols C,D,E,F,I,J,K,L,M,N,O,P,Z,AA,AB
  ReliableSoftware(.xlsx/.xls): header at row 1 (idx 0), data from row 2
                  cols B,E,F,K,L
  NiftyInvest     (.csv):       header at row 1 (idx 0), data from row 2
                  cols A,C
  MarketProfile   (.csv/.xlsx): header at row 1 (idx 0), data from row 2
                  cols B(stock key),G(VAH),H(POC),I(VAL)
"""

_SHAREKHAN_COLS  = [2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15, 25, 26, 27]
_RELIABLE_COLS   = [1, 4, 5, 10, 11]
_NIFTY_COLS      = [0, 2]
_MARKET_PROFILE_COLS = [1, 6, 7, 8]   # B=stock(key), G=VAH, H=POC, I=VAL

# Row index (0-based) of the header row for each broker
_SHAREKHAN_HEADER_ROW  = 7   # row 8
_RELIABLE_HEADER_ROW   = 0   # row 1
_NIFTY_HEADER_ROW      = 0   # row 1
_MARKET_PROFILE_HEADER_ROW = 0   # row 1


def count_rows(path: str, data_start_row: int = 1) -> int:
    """Return number of data rows. data_start_row is 0-based index of first data row."""
    try:
        if path.lower().endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            count = ws.max_row - data_start_row
            wb.close()
            return max(0, count)
        elif path.lower().endswith(".xls"):
            import xlrd
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            return max(0, ws.nrows - data_start_row)
        elif path.lower().endswith(".csv"):
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return max(0, sum(1 for _ in f) - data_start_row)
    except Exception:
        pass
    return 0


def count_rows_sharekhan(path: str) -> int:
    return count_rows(path, data_start_row=_SHAREKHAN_HEADER_ROW + 1)


def count_rows_reliable(path: str) -> int:
    return count_rows(path, data_start_row=_RELIABLE_HEADER_ROW + 1)


def count_rows_nifty(path: str) -> int:
    return count_rows(path, data_start_row=_NIFTY_HEADER_ROW + 1)


def count_rows_market_profile(path: str) -> int:
    return count_rows(path, data_start_row=_MARKET_PROFILE_HEADER_ROW + 1)


def _extract_cols(row: tuple | list, col_indices: list, ncols: int) -> list:
    return [row[i] if i < ncols else None for i in col_indices]


def _read_xlsx(path: str, col_indices: list, header_row_idx: int) -> tuple[list, list]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) <= header_row_idx:
        return [], []
    ncols = len(rows[header_row_idx]) if rows[header_row_idx] else 0
    headers = _extract_cols(rows[header_row_idx], col_indices, ncols)
    headers = [str(h) if h is not None else "" for h in headers]
    data = []
    for row in rows[header_row_idx + 1:]:
        nc = len(row) if row else 0
        data.append(_extract_cols(row, col_indices, nc))
    return headers, data


def _read_xls(path: str, col_indices: list, header_row_idx: int) -> tuple[list, list]:
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    if ws.nrows <= header_row_idx:
        return [], []
    headers = [str(ws.cell_value(header_row_idx, i)) if i < ws.ncols else ""
               for i in col_indices]
    data = []
    for r in range(header_row_idx + 1, ws.nrows):
        data.append([ws.cell_value(r, i) if i < ws.ncols else None for i in col_indices])
    return headers, data


def _read_csv(path: str, col_indices: list, header_row_idx: int) -> tuple[list, list]:
    import csv
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        rows = list(csv.reader(f))
    if len(rows) <= header_row_idx:
        return [], []
    ncols = len(rows[header_row_idx])
    headers = _extract_cols(rows[header_row_idx], col_indices, ncols)
    headers = [str(h) if h is not None else "" for h in headers]
    data = []
    for row in rows[header_row_idx + 1:]:
        nc = len(row)
        data.append(_extract_cols(row, col_indices, nc))
    return headers, data


def _read_file(path: str, col_indices: list, header_row_idx: int) -> tuple[list, list]:
    if path.lower().endswith(".xlsx"):
        return _read_xlsx(path, col_indices, header_row_idx)
    elif path.lower().endswith(".xls"):
        return _read_xls(path, col_indices, header_row_idx)
    elif path.lower().endswith(".csv"):
        return _read_csv(path, col_indices, header_row_idx)
    raise ValueError(f"Unsupported file type: {path}")


def read_sharekhan(path: str) -> tuple[list, list]:
    return _read_file(path, _SHAREKHAN_COLS, _SHAREKHAN_HEADER_ROW)


def read_reliable_software(path: str) -> tuple[list, list]:
    return _read_file(path, _RELIABLE_COLS, _RELIABLE_HEADER_ROW)


def read_reliable_software_date(path: str):
    """Returns the trade date from a ReliableSoftware export's first data
    row, column A (DataTime) — or None if unreadable/blank. Used only for
    the Data Import screen's freshness check; read_reliable_software()
    itself never reads this column (see _RELIABLE_COLS, which starts at B).
    """
    data_row = _RELIABLE_HEADER_ROW + 1
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) <= data_row or not rows[data_row]:
            return None
        return _xlsx_cell_date(rows[data_row][0])
    elif path.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        if ws.nrows <= data_row:
            return None
        return _xls_cell_date(ws.cell(data_row, 0), wb.datemode)
    raise ValueError(f"Unsupported file type: {path}")


def read_nifty_invest(path: str) -> tuple[list, list]:
    return _read_file(path, _NIFTY_COLS, _NIFTY_HEADER_ROW)


_NIFTY_SYMBOL_HEADER  = "Symbol"
_NIFTY_MAXPAIN_HEADER = "Max Pain"


def read_nifty_invest_multi(paths) -> tuple[list, list]:
    """Read one or more NiftyInvest exports and merge them into a single
    (headers, rows) pair, same output shape as read_nifty_invest.

    Unlike read_nifty_invest (fixed columns A/C), each file's Symbol and Max
    Pain columns are located BY HEADER NAME — different NiftyInvest exports
    may lay their columns out differently. A file missing either header is
    skipped rather than raising, so one bad file doesn't take down the
    live-read loop; user-facing rejection of a bad file happens earlier, at
    selection time (see screens.data_import.BrokerImportCard._validate_nifty_invest_file).

    Rows are merged by Symbol across all files, in the given path order —
    if the same Symbol appears in more than one file, the last file wins.

    `paths` may be a single path (str) or a list of paths.
    """
    if isinstance(paths, str):
        paths = [paths]

    combined: dict = {}
    for path in paths:
        try:
            headers, rows = read_external_import(path)
        except Exception:
            continue
        headers = [str(h).strip() if h is not None else "" for h in headers]
        if _NIFTY_SYMBOL_HEADER not in headers or _NIFTY_MAXPAIN_HEADER not in headers:
            continue
        sym_idx = headers.index(_NIFTY_SYMBOL_HEADER)
        mp_idx = headers.index(_NIFTY_MAXPAIN_HEADER)
        for row in rows:
            symbol = row[sym_idx] if sym_idx < len(row) else None
            if symbol is None or not str(symbol).strip():
                continue
            max_pain = row[mp_idx] if mp_idx < len(row) else None
            combined[str(symbol).strip()] = max_pain

    rows = [[symbol, max_pain] for symbol, max_pain in combined.items()]
    return [_NIFTY_SYMBOL_HEADER, _NIFTY_MAXPAIN_HEADER], rows


def read_market_profile(path: str) -> tuple[list, list]:
    return _read_file(path, _MARKET_PROFILE_COLS, _MARKET_PROFILE_HEADER_ROW)


def read_market_profile_date(path: str):
    """Returns the trade date from a MarketProfile export's first data row,
    column A (Date) — or None if unreadable/blank. Used only for the Data
    Import screen's freshness check; read_market_profile() itself never
    reads this column (see _MARKET_PROFILE_COLS, which starts at B).

    Unlike ReliableSoftware's DataTime, MarketProfile's Date often arrives
    as a plain "YYYY-MM-DD" string rather than a typed date cell (CSV has
    no cell types at all, and xlsx/xls exports vary by tool) — each branch
    tries the typed-date path first, then falls back to ISO string parsing.
    """
    data_row = _MARKET_PROFILE_HEADER_ROW + 1
    lower = path.lower()
    if lower.endswith(".csv"):
        import csv
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            rows = list(csv.reader(f))
        if len(rows) <= data_row or not rows[data_row]:
            return None
        return _parse_iso_date(rows[data_row][0])
    elif lower.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) <= data_row or not rows[data_row]:
            return None
        value = rows[data_row][0]
        return _xlsx_cell_date(value) or _parse_iso_date(value)
    elif lower.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        if ws.nrows <= data_row:
            return None
        cell = ws.cell(data_row, 0)
        return _xls_cell_date(cell, wb.datemode) or _parse_iso_date(cell.value)
    raise ValueError(f"Unsupported file type: {path}")


def read_external_import(path: str) -> tuple[list, list]:
    """Read all columns from an external file (xlsx/xls/csv). Header at row 1."""
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return [], []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        data = [list(r) for r in rows[1:]]
        return headers, data
    elif path.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return [], []
        headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
        data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(1, ws.nrows)]
        return headers, data
    elif path.lower().endswith(".csv"):
        import csv
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], []
        return rows[0], [list(r) for r in rows[1:]]
    raise ValueError(f"Unsupported file type: {path}")


def count_rows_external(path: str) -> int:
    return count_rows(path, data_start_row=1)


def _xls_cell_date(cell, datemode: int):
    """Return a datetime.date from an xlrd cell if it holds a date, else None."""
    import xlrd
    if cell.ctype != xlrd.XL_CELL_DATE:
        return None
    dt = xlrd.xldate_as_datetime(cell.value, datemode)
    return dt.date()


def _xlsx_cell_date(value):
    """Return a datetime.date from an openpyxl cell value if it's a date/datetime, else None."""
    import datetime as _dt
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    return None


def _parse_iso_date(value):
    """Return a datetime.date from a "YYYY-MM-DD"-ish string value, else None."""
    if value is None:
        return None
    from datetime import date as _date
    try:
        return _date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _drop_blank_scripname_rows(headers: list, data: list, row_dates: list) -> tuple[list, list]:
    """Excel sheets often carry stray formatting/leftover values past the
    real data range — xlrd/openpyxl report these as trailing rows in the
    sheet's used range, blank except for the odd artifact value. A row with
    no ScripName can't be attributed to any stock, so it's dropped here
    rather than uploaded later as an empty symbol.
    """
    if "ScripName" not in headers:
        return data, row_dates
    scrip_idx = headers.index("ScripName")
    kept_data, kept_dates = [], []
    for i, row in enumerate(data):
        value = row[scrip_idx] if scrip_idx < len(row) else None
        if value is None or not str(value).strip():
            continue
        kept_data.append(row)
        kept_dates.append(row_dates[i] if i < len(row_dates) else None)
    return kept_data, kept_dates


def read_historic_sheet(path: str) -> tuple[list, list, list]:
    """
    Read all columns from a historic-upload file (xlsx/xls/csv), no fixed
    template, and additionally extract the trade date from a "DataTime"
    column (time part dropped) for every row so the caller can validate the
    sheet's dates against the user-picked upload date.

    Rows with a blank ScripName are dropped (see _drop_blank_scripname_rows)
    — a stray trailing blank row in the sheet is not a real stock row.

    Returns (headers, rows, row_dates) — row_dates[i] is a datetime.date or
    None if that row's DataTime cell wasn't a parseable date, or [] filled
    with None for every row if there's no "DataTime" column at all.
    """
    from datetime import date as _date

    lower = path.lower()
    if lower.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not raw_rows:
            return [], [], []
        headers = [str(h) if h is not None else "" for h in raw_rows[0]]
        data = [list(r) for r in raw_rows[1:]]
        date_col = headers.index("DataTime") if "DataTime" in headers else None
        row_dates = [
            _xlsx_cell_date(row[date_col]) if date_col is not None and date_col < len(row) else None
            for row in data
        ]
        data, row_dates = _drop_blank_scripname_rows(headers, data, row_dates)
        return headers, data, row_dates

    elif lower.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return [], [], []
        headers = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
        data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(1, ws.nrows)]
        date_col = headers.index("DataTime") if "DataTime" in headers else None
        row_dates = []
        for r in range(1, ws.nrows):
            if date_col is None:
                row_dates.append(None)
            else:
                row_dates.append(_xls_cell_date(ws.cell(r, date_col), wb.datemode))
        data, row_dates = _drop_blank_scripname_rows(headers, data, row_dates)
        return headers, data, row_dates

    elif lower.endswith(".csv"):
        headers, data = read_external_import(path)
        date_col = headers.index("DataTime") if "DataTime" in headers else None
        row_dates = []
        for row in data:
            parsed = None
            if date_col is not None and date_col < len(row) and row[date_col]:
                try:
                    parsed = _date.fromisoformat(str(row[date_col])[:10])
                except ValueError:
                    parsed = None
            row_dates.append(parsed)
        data, row_dates = _drop_blank_scripname_rows(headers, data, row_dates)
        return headers, data, row_dates

    raise ValueError(f"Unsupported file type: {path}")
