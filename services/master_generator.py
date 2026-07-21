"""
Merge data from three broker sources into a single sheet.

Join strategy:
  - Sharekhan Column C (Scrip Name) is the primary key
  - ReliableSoftware Column B (ScripName) is a foreign key — resolved to
    symbol via the Script Name mapping table before matching
  - NiftyInvest Symbol (located by header name, not a fixed column — see
    services.file_reader.read_nifty_invest_multi; multiple NiftyInvest files
    are supported and merged, last file wins on a duplicate Symbol) is a
    foreign key — matched directly
  - All rows from Sharekhan are kept; missing data from other sources = blank

Column positions within each read result (0-based within the extracted cols):
  Sharekhan  : [0]=Scrip Name, [1]=Lot Size, [2]=% Change, [3]=Current,
               [4]=Open, [5]=High, [6]=Low, [7]=Close, [8]=Avg Rate,
               [9]=OI Difference Percentage, [10]=P.High, [11]=P.Low,
               [12]=Qty, [13]=P.Quantity, [14]=TurnOver
  ReliableSoft: [0]=ScripName (FK — omitted from output), [1]=callstrikehighestoi,
               [2]=Callstrikewithsecondhighestoi, [3]=PutStrikeWithsecondHighestOI,
               [4]=TodayPutHighestStrike
  NiftyInvest: [0]=Symbol (FK — omitted from output), [1]=Max Pain
"""

import os
from datetime import date, timedelta
from services.file_reader import (read_sharekhan, read_reliable_software,
                                  read_nifty_invest_multi, read_external_import,
                                  read_market_profile)


def last_tuesday_of_month(ref_date: date = None) -> date:
    """Return the last Tuesday of the month for the given date."""
    if ref_date is None:
        ref_date = date.today()
    year, month = ref_date.year, ref_date.month
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    d = last_day
    while d.weekday() != 1:  # 1 = Tuesday
        d -= timedelta(days=1)
    return d


# Indices within extracted result rows for each broker's primary/foreign key
_SK_PK_IDX  = 0   # Sharekhan  → Scrip Name
_RS_FK_IDX  = 0   # ReliableSoftware → ScripName (full name, needs mapping)
_NI_FK_IDX  = 0   # NiftyInvest → Symbol
_MP_FK_IDX  = 0   # MarketProfile → stock (already a symbol, direct match)

# Indices of data columns to include in output (foreign key col excluded)
_RS_DATA_INDICES = [1, 2, 3, 4]   # skip index 0 (ScripName FK)
_NI_DATA_INDICES = [1]             # skip index 0 (Symbol FK)
_MP_DATA_INDICES = [1, 2, 3]       # skip index 0 (stock FK): VAH, POC, VAL


def _build_script_name_lookup(script_name_data: list) -> dict:
    """
    Build a dict from full name → symbol using the Script Name config table.
    script_name_data is a list of (full_name, symbol) tuples.
    Both keys are stored normalised (stripped, lowercased) for robust matching.
    """
    return {str(full).strip().lower(): str(sym).strip()
            for full, sym in script_name_data}


def _normalise(value) -> str:
    return str(value).strip() if value is not None else ""


def _strip_rolling_suffix(name: str) -> str:
    """
    Strip a trailing rolling-expiry suffix from a ReliableSoftware ScripName.

    The suffix looks like ``.rolling.12D`` (or .11D, .10D, …) appended to the
    company name. We cut at the ".rolling" marker itself (case-insensitive)
    rather than assuming it's always exactly the last two dot-separated
    segments — a fixed-position rsplit silently leaves a ".rolling" fragment
    behind whenever the suffix isn't the very last two segments (e.g. an
    extra trailing segment after the expiry code), which then fails to match
    the Script Name config and leaks into the saved display name:

        "ABB LTD.rolling.12D"          -> "ABB LTD"
        "ADANIENT.rolling.11D"         -> "ADANIENT"
        "BIOCON LIMITED..rolling.12D"  -> "BIOCON LIMITED."

    Names without a rolling suffix are returned unchanged.
    """
    s = str(name).strip() if name is not None else ""
    idx = s.lower().find(".rolling")
    if idx == -1:
        return s
    return s[:idx]


def generate_master(
    sharekhan_path: str,
    reliable_path: str,
    nifty_paths,   # a single path (str) or a list of paths — see read_nifty_invest_multi
    output_path: str,
    script_name_data: list,   # [(full_name, symbol), ...] from config tab 2
    expiry_date: date = None,  # expiry date to strip from Sharekhan Scrip Names
    external_path: str = None,
    market_profile_path: str = None,
) -> None:
    # --- Read all sources ---
    sk_headers, sk_rows = read_sharekhan(sharekhan_path)
    rs_headers, rs_rows = read_reliable_software(reliable_path)
    ni_headers, ni_rows = read_nifty_invest_multi(nifty_paths)
    if external_path:
        ext_headers, ext_rows = read_external_import(external_path)
    else:
        ext_headers, ext_rows = [], []
    if market_profile_path:
        mp_headers, mp_rows = read_market_profile(market_profile_path)
    else:
        mp_headers, mp_rows = [], []

    # --- Strip expiry date suffix from Sharekhan Scrip Names ---
    if expiry_date is not None:
        expiry_str = expiry_date.strftime("%d-%b-%Y").upper()
        for sk_row in sk_rows:
            scrip = _normalise(sk_row[_SK_PK_IDX])
            if scrip.upper().endswith(expiry_str):
                sk_row[_SK_PK_IDX] = scrip[:-len(expiry_str)].strip()

    # --- Build Script Name lookup: full_name_lower → symbol ---
    name_to_symbol = _build_script_name_lookup(script_name_data)

    # --- Build lookup maps keyed by normalised symbol ---
    # ReliableSoftware: resolve full name → symbol first
    rs_lookup: dict[str, list] = {}
    for row in rs_rows:
        full_name = _strip_rolling_suffix(_normalise(row[_RS_FK_IDX]))
        symbol = name_to_symbol.get(full_name.lower())
        if symbol:
            key = _normalise(symbol).upper()
            rs_lookup[key] = row

    # NiftyInvest: symbol is already the key
    ni_lookup: dict[str, list] = {}
    for row in ni_rows:
        key = _normalise(row[_NI_FK_IDX]).upper()
        ni_lookup[key] = row

    # External import: same shape as ReliableSoftware — column B (index 1) is
    # the join key (full name + rolling suffix → symbol via config), column C
    # onward (index 2+) are the data columns.
    ext_data_indices = list(range(2, len(ext_headers))) if len(ext_headers) > 2 else []
    ext_lookup: dict = {}
    if ext_rows and ext_headers:
        for row in ext_rows:
            if not row or len(row) < 2:
                continue
            full_name = _strip_rolling_suffix(_normalise(row[1]))
            symbol = name_to_symbol.get(full_name.lower())
            if symbol:
                ext_lookup[_normalise(symbol).upper()] = row

    # MarketProfile: stock (index 0) is already a symbol → direct match.
    mp_lookup: dict = {}
    for row in mp_rows:
        key = _normalise(row[_MP_FK_IDX]).upper()
        if key:
            mp_lookup[key] = row

    # --- Build merged headers ---
    out_headers = list(sk_headers)
    for i in _RS_DATA_INDICES:
        out_headers.append(rs_headers[i] if i < len(rs_headers) else "")
    for i in _NI_DATA_INDICES:
        out_headers.append(ni_headers[i] if i < len(ni_headers) else "")
    for i in ext_data_indices:
        out_headers.append(ext_headers[i] if i < len(ext_headers) else "")
    for i in _MP_DATA_INDICES:
        out_headers.append(mp_headers[i] if i < len(mp_headers) else "")

    # --- Build merged rows ---
    merged_rows = []
    for sk_row in sk_rows:
        pk = _normalise(sk_row[_SK_PK_IDX]).upper()

        out_row = list(sk_row)

        rs_row = rs_lookup.get(pk)
        for i in _RS_DATA_INDICES:
            out_row.append(rs_row[i] if rs_row and i < len(rs_row) else None)

        ni_row = ni_lookup.get(pk)
        for i in _NI_DATA_INDICES:
            out_row.append(ni_row[i] if ni_row and i < len(ni_row) else None)

        ext_row = ext_lookup.get(pk)
        for i in ext_data_indices:
            out_row.append(ext_row[i] if ext_row and i < len(ext_row) else None)

        mp_row = mp_lookup.get(pk)
        for i in _MP_DATA_INDICES:
            out_row.append(mp_row[i] if mp_row and i < len(mp_row) else None)

        merged_rows.append(out_row)

    _write_xlsx(output_path, out_headers, merged_rows)


def _write_xlsx(output_path: str, headers: list, rows: list) -> None:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    # Build the workbook fully in memory first, then write all bytes to the
    # existing file in one shot. This preserves the inode so Excel/Numbers
    # (which watch the inode via FSEvents) fire their external-change
    # notification and reload the file while it is open.
    buf = io.BytesIO()
    wb.save(buf)
    with open(output_path, "wb") as f:
        f.write(buf.getvalue())
