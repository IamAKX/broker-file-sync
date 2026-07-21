"""Tests for services.file_reader — historic-upload sheet reading."""
import sys
import os
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from services.file_reader import (
    _drop_blank_scripname_rows,
    read_historic_sheet,
    read_market_profile_date,
    read_reliable_software_date,
    read_nifty_invest_multi,
)


def test_drop_blank_scripname_rows_removes_blank_rows():
    headers = ["ScripName", "Open"]
    data = [["INFY", 100], ["", 0], ["TCS", 200], [None, 0]]
    row_dates = [None, None, None, None]
    kept_data, kept_dates = _drop_blank_scripname_rows(headers, data, row_dates)
    assert kept_data == [["INFY", 100], ["TCS", 200]]
    assert len(kept_dates) == 2


def test_drop_blank_scripname_rows_noop_when_no_scripname_column():
    headers = ["Symbol", "Open"]
    data = [["INFY", 100], ["", 0]]
    row_dates = [None, None]
    kept_data, kept_dates = _drop_blank_scripname_rows(headers, data, row_dates)
    assert kept_data == data
    assert kept_dates == row_dates


def test_read_historic_sheet_xlsx_drops_trailing_blank_rows(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DataTime", "ScripName", "Open"])
    ws.append([46149, "INFY", 100])
    ws.append([46149, "TCS", 200])
    # Trailing rows with a blank ScripName but a stray leftover value —
    # mirrors the real-world Excel quirk this fix targets.
    ws.append(["", "", ""])
    ws.append(["", "", 5749])
    path = str(tmp_path / "historic.xlsx")
    wb.save(path)

    headers, rows, row_dates = read_historic_sheet(path)

    assert headers == ["DataTime", "ScripName", "Open"]
    assert len(rows) == 2
    assert len(row_dates) == 2
    assert [r[1] for r in rows] == ["INFY", "TCS"]


def test_read_reliable_software_date_reads_first_data_row_column_a(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DataTime", "ScripName", "DiffPcnt"])
    ws.append([datetime.datetime(2026, 6, 18), "INFY.rolling.12D", 1.5])
    ws.append([datetime.datetime(2026, 6, 18), "TCS.rolling.12D", 2.0])
    path = str(tmp_path / "reliable.xlsx")
    wb.save(path)

    assert read_reliable_software_date(path) == datetime.date(2026, 6, 18)


def test_read_reliable_software_date_returns_none_when_column_a_blank(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DataTime", "ScripName", "DiffPcnt"])
    ws.append([None, "INFY.rolling.12D", 1.5])
    path = str(tmp_path / "reliable.xlsx")
    wb.save(path)

    assert read_reliable_software_date(path) is None


def test_read_reliable_software_date_returns_none_when_no_data_rows(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["DataTime", "ScripName", "DiffPcnt"])
    path = str(tmp_path / "reliable.xlsx")
    wb.save(path)

    assert read_reliable_software_date(path) is None


def test_read_market_profile_date_reads_iso_string_from_csv(tmp_path):
    path = tmp_path / "marketprofile.csv"
    path.write_text(
        "Date,stock,Open,High,Low,VAH,POC,VAL\n"
        "2026-06-29,360ONE,1079.8,1097.0,1070.4,1094.0,1088.0,1086.0\n"
        "2026-06-29,ABB,7000.0,7268.0,6887.5,7117.5,7007.5,6967.5\n"
    )
    assert read_market_profile_date(str(path)) == datetime.date(2026, 6, 29)


def test_read_market_profile_date_returns_none_when_blank_csv(tmp_path):
    path = tmp_path / "marketprofile.csv"
    path.write_text("Date,stock,Open\n,360ONE,1079.8\n")
    assert read_market_profile_date(str(path)) is None


def test_read_market_profile_date_returns_none_when_no_data_rows_csv(tmp_path):
    path = tmp_path / "marketprofile.csv"
    path.write_text("Date,stock,Open\n")
    assert read_market_profile_date(str(path)) is None


def test_read_market_profile_date_reads_typed_date_from_xlsx(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "stock", "Open"])
    ws.append([datetime.datetime(2026, 6, 29), "360ONE", 1079.8])
    path = str(tmp_path / "marketprofile.xlsx")
    wb.save(path)

    assert read_market_profile_date(path) == datetime.date(2026, 6, 29)


def test_read_market_profile_date_reads_iso_string_from_xlsx(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "stock", "Open"])
    ws.append(["2026-06-29", "360ONE", 1079.8])
    path = str(tmp_path / "marketprofile.xlsx")
    wb.save(path)

    assert read_market_profile_date(path) == datetime.date(2026, 6, 29)


# ── read_nifty_invest_multi ─────────────────────────────────────────────────

def _write_csv(tmp_path, name, header_row, rows):
    path = tmp_path / name
    lines = [",".join(header_row)] + [",".join(str(c) for c in r) for r in rows]
    path.write_text("\n".join(lines) + "\n")
    return str(path)


def test_read_nifty_invest_multi_locates_columns_by_header_name_in_different_positions(tmp_path):
    # File A: Symbol first, Max Pain second. File B: reversed, plus an extra column.
    f1 = _write_csv(tmp_path, "a.csv", ["Symbol", "Max Pain"], [["INFY", 1800]])
    f2 = _write_csv(tmp_path, "b.csv", ["Max Pain", "Symbol", "Extra"], [[3500, "TCS", "x"]])

    headers, rows = read_nifty_invest_multi([f1, f2])
    assert headers == ["Symbol", "Max Pain"]
    rows_by_symbol = {r[0]: r[1] for r in rows}
    assert rows_by_symbol["INFY"] == "1800"
    assert rows_by_symbol["TCS"] == "3500"


def test_read_nifty_invest_multi_last_file_wins_on_duplicate_symbol(tmp_path):
    f1 = _write_csv(tmp_path, "a.csv", ["Symbol", "Max Pain"], [["INFY", 1800]])
    f2 = _write_csv(tmp_path, "b.csv", ["Symbol", "Max Pain"], [["INFY", 1900]])

    headers, rows = read_nifty_invest_multi([f1, f2])
    rows_by_symbol = {r[0]: r[1] for r in rows}
    assert rows_by_symbol["INFY"] == "1900"


def test_read_nifty_invest_multi_skips_file_missing_required_header(tmp_path):
    f1 = _write_csv(tmp_path, "a.csv", ["Symbol", "Max Pain"], [["INFY", 1800]])
    f2 = _write_csv(tmp_path, "b.csv", ["Symbol", "SomethingElse"], [["TCS", 3500]])

    headers, rows = read_nifty_invest_multi([f1, f2])
    rows_by_symbol = {r[0]: r[1] for r in rows}
    assert rows_by_symbol == {"INFY": "1800"}


def test_read_nifty_invest_multi_accepts_a_single_string_path(tmp_path):
    f1 = _write_csv(tmp_path, "a.csv", ["Symbol", "Max Pain"], [["INFY", 1800]])
    headers, rows = read_nifty_invest_multi(f1)
    assert rows == [["INFY", "1800"]]


def test_read_nifty_invest_multi_empty_list_returns_empty_rows():
    headers, rows = read_nifty_invest_multi([])
    assert headers == ["Symbol", "Max Pain"]
    assert rows == []


def test_read_nifty_invest_multi_skips_blank_symbol_rows(tmp_path):
    f1 = _write_csv(tmp_path, "a.csv", ["Symbol", "Max Pain"], [["INFY", 1800], ["", 999]])
    headers, rows = read_nifty_invest_multi([f1])
    assert rows == [["INFY", "1800"]]
