"""Tests for the Market Profile (5th file) import source."""
import csv
import sys
import os
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


@pytest.fixture(scope="module")
def qapp():
    from PySide6.QtWidgets import QApplication
    return QApplication.instance() or QApplication(sys.argv)


def _make_csv(tmp_path, name, rows):
    f = tmp_path / name
    with open(f, "w", newline="") as fh:
        w = csv.writer(fh)
        for r in rows:
            w.writerow(r)
    return str(f)


_HEADER = ["Date", "stock", "Open", "High", "Low", "Day's Range",
           "VAH", "POC", "VAL", "IB Range", "Close", "IB High", "IB Low", "TPO at POC"]


# ── file_reader ─────────────────────────────────────────────────────────────

def test_read_market_profile_extracts_b_g_h_i(tmp_path):
    from services.file_reader import read_market_profile
    f = _make_csv(tmp_path, "mp.csv", [
        _HEADER,
        ["2026-06-29", "360ONE", "1079.8", "1097.0", "1070.4", "26.6",
         "1094.0", "1088.0", "1086.0", "24.0", "1084.0", "1095.0", "1071.0", "11.0"],
    ])
    headers, data = read_market_profile(f)
    # B=stock, G=VAH, H=POC, I=VAL
    assert headers == ["stock", "VAH", "POC", "VAL"]
    assert data[0] == ["360ONE", "1094.0", "1088.0", "1086.0"]


def test_count_rows_market_profile(tmp_path):
    from services.file_reader import count_rows_market_profile
    f = _make_csv(tmp_path, "mp.csv", [_HEADER, ["d", "A"], ["d", "B"], ["d", "C"]])
    assert count_rows_market_profile(f) == 3


# ── live_merge ──────────────────────────────────────────────────────────────

def _stub_other_readers(monkeypatch):
    monkeypatch.setattr(
        "services.live_merge.LiveDataReader._read_sharekhan",
        lambda self: (["Scrip Name", "Current"], [["360ONE", 1.0], ["ABB", 2.0]]),
    )
    monkeypatch.setattr(
        "services.file_reader.read_reliable_software",
        lambda path: (["ScripName", "callstrikehighestoi"], []),
    )
    monkeypatch.setattr(
        "services.file_reader.read_nifty_invest",
        lambda path: (["Symbol", "Max Pain"], []),
    )


def test_live_merge_includes_market_profile_columns(tmp_path, monkeypatch):
    from services.live_merge import LiveDataReader
    _stub_other_readers(monkeypatch)
    mp_file = _make_csv(tmp_path, "mp.csv", [
        _HEADER,
        ["2026-06-29", "360ONE", "x", "x", "x", "x", "1094", "1088", "1086",
         "x", "x", "x", "x", "x"],
        ["2026-06-29", "ABB", "x", "x", "x", "x", "7117", "7007", "6967",
         "x", "x", "x", "x", "x"],
    ])
    reader = LiveDataReader("sk.xlsx", "rs.xlsx", "ni.csv", [],
                            market_profile_path=mp_file)
    reader._read_slow_sources(force=True)
    headers, data = reader.read_merged(force_slow=True)

    for col in ("VAH", "POC", "VAL"):
        assert col in headers
    rows = {r[0]: r for r in data}
    assert rows["360ONE"][headers.index("VAH")] == "1094"
    assert rows["360ONE"][headers.index("POC")] == "1088"
    assert rows["360ONE"][headers.index("VAL")] == "1086"
    assert rows["ABB"][headers.index("VAH")] == "7117"


def test_live_merge_market_profile_missing_scrip_is_none(tmp_path, monkeypatch):
    from services.live_merge import LiveDataReader
    _stub_other_readers(monkeypatch)
    # Only 360ONE present — ABB should get None for MP columns.
    mp_file = _make_csv(tmp_path, "mp.csv", [
        _HEADER,
        ["2026-06-29", "360ONE", "x", "x", "x", "x", "1094", "1088", "1086",
         "x", "x", "x", "x", "x"],
    ])
    reader = LiveDataReader("sk.xlsx", "rs.xlsx", "ni.csv", [],
                            market_profile_path=mp_file)
    reader._read_slow_sources(force=True)
    headers, data = reader.read_merged(force_slow=True)
    rows = {r[0]: r for r in data}
    assert rows["ABB"][headers.index("VAH")] is None


def test_live_merge_without_market_profile_unchanged(tmp_path, monkeypatch):
    from services.live_merge import LiveDataReader
    _stub_other_readers(monkeypatch)
    reader = LiveDataReader("sk.xlsx", "rs.xlsx", "ni.csv", [])
    reader._read_slow_sources(force=True)
    headers, data = reader.read_merged(force_slow=True)
    assert "VAH" not in headers
    assert len(data) == 2


# ── master_generator ────────────────────────────────────────────────────────

def test_generate_master_includes_market_profile(tmp_path, monkeypatch):
    pytest.importorskip("openpyxl")
    import openpyxl
    from services import master_generator as mg

    monkeypatch.setattr(mg, "read_sharekhan",
                        lambda p: (["Scrip Name", "Current"], [["360ONE", "1.0"]]))
    monkeypatch.setattr(mg, "read_reliable_software", lambda p: (["ScripName"], []))
    monkeypatch.setattr(mg, "read_nifty_invest_multi", lambda p: (["Symbol"], []))
    mp_file = _make_csv(tmp_path, "mp.csv", [
        _HEADER,
        ["2026-06-29", "360ONE", "x", "x", "x", "x", "1094", "1088", "1086",
         "x", "x", "x", "x", "x"],
    ])
    out = str(tmp_path / "master.xlsx")
    mg.generate_master("sk", "rs", "ni", out, [], market_profile_path=mp_file)

    ws = openpyxl.load_workbook(out).active
    header_row = [c.value for c in ws[1]]
    assert "VAH" in header_row and "POC" in header_row and "VAL" in header_row
    vah_col = header_row.index("VAH")
    assert ws[2][vah_col].value == "1094"


# ── UI wiring ─────────────────────────────────────────────────────────────────

def test_data_import_has_market_profile_card(qapp):
    from app import AppController
    from screens.data_import import DataImportScreen
    screen = DataImportScreen(AppController(qapp))
    assert "MarketProfile" in screen._cards


def test_market_profile_in_required_brokers():
    from screens.data_import import DataImportScreen
    assert "MarketProfile" in DataImportScreen._REQUIRED_BROKERS


def test_market_profile_in_brokers_list_with_pink():
    from screens.data_import import BROKERS
    mp = next(b for b in BROKERS if b[0] == "MarketProfile")
    assert mp[1] == "status_pink"
    assert ".csv" in mp[3] and ".xlsx" in mp[3]


def test_dashboard_lists_market_profile(qapp):
    from screens.dashboard import BROKER_COLORS
    names = [n for n, _ in BROKER_COLORS]
    assert "MarketProfile" in names
    assert ("MarketProfile", "status_pink") in BROKER_COLORS


def test_sidebar_lists_market_profile_pink(qapp):
    from components.sidebar import BROKERS
    assert ("MarketProfile", "status_pink") in BROKERS


def test_theme_has_status_pink():
    from theme import DARK, LIGHT
    assert "status_pink" in DARK
    assert "status_pink" in LIGHT


def test_live_viewer_accepts_market_profile_path(qapp, tmp_path, monkeypatch):
    from services import strategy_store as store
    monkeypatch.setattr(store, "_STORE_FILE", str(tmp_path / "s.json"))
    from screens.live_viewer import LiveViewerWindow
    lmv = LiveViewerWindow("", "", "", [],
                           market_profile_path=str(tmp_path / "mp.csv"))
    assert lmv._market_profile_path is not None
