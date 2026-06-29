import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest


def test_apply_rename_overrides_only_mapped():
    from services.lmv_export import apply_rename
    headers = ["Current", "High", "Low"]
    out = apply_rename(headers, {"Current": "LTP", "Low": "Bottom"})
    assert out == ["LTP", "High", "Bottom"]


def test_apply_rename_empty_map_is_identity():
    from services.lmv_export import apply_rename
    headers = ["A", "B"]
    assert apply_rename(headers, {}) == ["A", "B"]


def test_export_xlsx_writes_renamed_headers(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from services.lmv_export import export_xlsx
    out = tmp_path / "export.xlsx"
    headers = ["Current", "High"]
    rows = [["100", "110"], ["200", "210"]]
    export_xlsx(str(out), headers, rows, rename_map={"Current": "LTP"})

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["LTP", "High"]
    assert [c.value for c in ws[2]] == ["100", "110"]
    assert ws.max_row == 3  # header + 2 rows


def test_export_xlsx_uses_config_store_map_by_default(tmp_path, monkeypatch):
    pytest.importorskip("openpyxl")
    import openpyxl
    from services import config_store
    monkeypatch.setattr(config_store, "_STORE_FILE", str(tmp_path / "config_data.json"))
    config_store.save_tab("main_column_name", [("Current", "LTP")])

    from services.lmv_export import export_xlsx
    out = tmp_path / "e.xlsx"
    export_xlsx(str(out), ["Current", "Open"], [["1", "2"]])
    ws = openpyxl.load_workbook(str(out)).active
    assert [c.value for c in ws[1]] == ["LTP", "Open"]
